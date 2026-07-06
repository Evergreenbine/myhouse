# -*- coding: utf-8 -*-
"""AI Tool Calls —— 定义和执行本地工具函数"""
import json
import os
from datetime import date, datetime, timedelta

# ============ Tool 定义 ============

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "query_attendance",
            "description": (
                "【万能考勤工具】查询任意日期范围的完整考勤数据，支持跨月查询。"
                "返回每日明细（打卡次数、加班小时、加班费、最早最晚打卡时间、是否漏打卡）"
                "和预计算汇总（加班最多日、加班费最多日、最晚打卡日等）。\n"
                "使用场景：用户问任何与考勤/打卡/加班/漏打卡/加班费有关的问题，都用这个。\n"
                "示例触发词：上班了吗、打卡记录、考勤、加班、加班费、漏打卡、最晚下班、加班最多哪天、加班统计"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date_from": {"type": "string", "description": "起始日期 YYYY-MM-DD"},
                    "date_to": {"type": "string", "description": "结束日期 YYYY-MM-DD"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_today_info",
            "description": (
                "【今日概览】获取今天的所有信息：打卡状态、加班情况、喝水进度、待办数量、天气。"
                "用户问今天的情况统一用这个。\n"
                "触发词：今天、今日、今天怎么样、今天状态、今天上班、今天打卡"
            ),
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "do_punch_card",
            "description": (
                "【补卡】为指定日期补打卡记录。先必须用 query_attendance 确认当天漏打卡再补。\n"
                "触发词：补卡、补打卡、补上打卡、忘记打卡、缺打卡"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "日期 YYYY-MM-DD"},
                    "time": {"type": "string", "description": "时间 HH:MM:SS"},
                    "remark": {"type": "string", "description": "打卡位置，如'公司大门打卡'"},
                },
                "required": ["date", "time"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_car_abnormal",
            "description": (
                "【车辆异常】查询车辆出入异常记录。\n"
                "触发词：车辆、车辆异常、车辆出入、车、停车、出入异常"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date_from": {"type": "string", "description": "起始日期 YYYY-MM-DD"},
                    "date_to": {"type": "string", "description": "结束日期 YYYY-MM-DD"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_water_status",
            "description": (
                "【喝水】查看喝水进度。\n"
                "触发词：喝水、水、喝了几杯、饮水"
            ),
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "query_todos",
            "description": (
                "【待办】查询待办事项。\n"
                "触发词：待办、todo、任务"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "only_undone": {"type": "boolean", "description": "仅未完成，默认true"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "do_water_punch",
            "description": (
                "喝水打卡，记录今天喝了第几杯水。先调用 get_water_status 查看当前进度再打卡。\n"
                "触发词：喝水、打水卡、喝水打卡、记一杯水"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "index": {
                        "type": "integer",
                        "description": "已喝杯数的索引（0=第1杯，7=第8杯）。检查当前进度后传入下一个杯数。",
                    },
                },
                "required": ["index"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "add_todo",
            "description": (
                "添加一条待办事项。\n"
                "触发词：添加待办、加个待办、提醒我、记一下、帮我记"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "日期 YYYY-MM-DD，默认今天"},
                    "time": {"type": "string", "description": "时间 HH:MM，默认当前时间"},
                    "content": {"type": "string", "description": "待办内容"},
                },
                "required": ["content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "toggle_todo",
            "description": (
                "切换待办事项的完成状态（已完成\u2194未完成）。\n"
                "触发词：完成待办、标记完成、取消完成、勾掉"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "待办日期 YYYY-MM-DD"},
                    "id": {"type": "string", "description": "待办ID，从 query_todos 获得"},
                },
                "required": ["date", "id"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "query_car_records",
            "description": (
                "【车辆出入记录】查询车辆/车牌的具体出入时间点（几点进、几点出）。\n"
                "返回所有出入记录列表，不是只返回异常记录。\n"
                "触发词：车辆出入记录、出入时间、进出记录、车进出"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date_from": {"type": "string", "description": "起始日期 YYYY-MM-DD"},
                    "date_to": {"type": "string", "description": "结束日期 YYYY-MM-DD"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "export_attendance_excel",
            "description": (
                "导出考勤报表为Excel文件。生成包含每日明细和汇总的.xlsx文件。\n"
                "触发词：导出Excel、导出报表、生成Excel、下载考勤、生成报表"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date_from": {"type": "string", "description": "起始日期 YYYY-MM-DD，默认本月1号"},
                    "date_to": {"type": "string", "description": "结束日期 YYYY-MM-DD，默认今天"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_todo",
            "description": (
                "删除一条待办事项。\n"
                "触发词：删除待办、移除待办、删掉"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "待办日期 YYYY-MM-DD"},
                    "id": {"type": "string", "description": "待办ID，从 query_todos 获得"},
                },
                "required": ["date", "id"],
            },
        },
    },

]


# ============ Tool 执行器 ============

class ToolExecutor:
    """执行AI请求的Tool Call，返回结果"""
    _instance = None

    def __init__(self, app_ref=None):
        self.app = app_ref  # OvertimeApp 实例

    def set_app(self, app):
        self.app = app

    @staticmethod
    def _get_empno():
        try:
            from local_db import load_app_user
            local = load_app_user()
            if local and local.get('empno'):
                return local.get('empno')
        except:
            pass
        try:
            from user_auth import load_user_config
            cfg = load_user_config()
            return cfg.empno if cfg else ''
        except:
            return ''

    def execute(self, tool_name, arguments):
        """执行单个工具调用，返回结果字符串"""
        try:
            if tool_name == "query_attendance":
                return self._query_attendance(arguments)
            elif tool_name == "do_punch_card":
                return self._do_punch_card(arguments)
            elif tool_name == "get_car_abnormal":
                return self._get_car_abnormal(arguments)
            elif tool_name == "query_todos":
                return self._query_todos(arguments)
            elif tool_name == "get_water_status":
                return self._get_water_status(arguments)
            elif tool_name == "get_today_info":
                return self._get_today_info(arguments)
            elif tool_name == "get_missed_punch_dates":
                # 兼容旧调用，重定向到 query_attendance
                return self._query_attendance({})
            elif tool_name == "get_overtime_stats":
                # 兼容旧调用
                return self._query_attendance({})
            elif tool_name == "query_car_records":
                return self._get_car_abnormal(arguments)
            elif tool_name == "export_attendance_excel":
                return self._export_attendance_excel(arguments)
            elif tool_name == "add_todo":
                return self._add_todo(arguments)
            elif tool_name == "toggle_todo":
                return self._toggle_todo(arguments)
            elif tool_name == "delete_todo":
                return self._delete_todo(arguments)
            elif tool_name == "do_water_punch":
                return self._do_water_punch(arguments)
            else:
                return f"未知工具: {tool_name}"
        except Exception as e:
            return f"工具执行出错: {str(e)[:200]}"

    # ---- 各工具实现 ----

    def _query_attendance(self, args):
        today = date.today()
        df = args.get("date_from", today.replace(day=1).isoformat())
        dt = args.get("date_to", today.isoformat())
        try:
            from user_auth import load_user_config
            from database import query_work_records_range
            cfg = load_user_config()
            emp = self._get_empno()
            if not emp:
                return "请先登录"
            s, e = date.fromisoformat(df), date.fromisoformat(dt)
            recs = query_work_records_range(emp, df, dt)
            from attendance import DayAnalysis
            results = []
            d = s
            while d <= e:
                ds = d.isoformat()
                day_recs = recs.get(ds, [])
                a = DayAnalysis(d, day_recs, cfg)
                punch_times = []
                for r in day_recs:
                    wt = r.get("worktime")
                    if wt:
                        punch_times.append(wt.strftime("%H:%M:%S") if hasattr(wt, "strftime") else str(wt)[-8:])
                punch_times.sort()
                results.append({
                    "日期": ds, "类型": a.day_type_label,
                    "加班小时": round(a.overtime_hours, 1),
                    "加班费": round(a.overtime_pay, 2),
                    "打卡次数": f"{a.card_count}/{a.required_cards}",
                    "漏打卡": a.missed,
                    "最早打卡": punch_times[0] if punch_times else "",
                    "最晚打卡": punch_times[-1] if punch_times else "",
                })
                d += timedelta(days=1)
            summary = {
                "查询范围": f"{df} ~ {dt}",
                "总天数": len(results),
                "加班天数": sum(1 for r in results if r["加班小时"] > 0),
                "漏打卡天数": sum(1 for r in results if r["漏打卡"]),
            }
            max_ot = max(results, key=lambda r: r["加班小时"]) if results else None
            if max_ot and max_ot["加班小时"] > 0:
                summary["加班最多日"] = {"日期": max_ot["日期"], "小时": max_ot["加班小时"]}
            max_pay = max(results, key=lambda r: r["加班费"]) if results else None
            if max_pay and max_pay["加班费"] > 0:
                summary["加班费最多日"] = {"日期": max_pay["日期"], "金额": max_pay["加班费"]}
            has_punch = [r for r in results if r["最晚打卡"]]
            if has_punch:
                latest = max(has_punch, key=lambda r: r["最晚打卡"])
                summary["最晚打卡日"] = {"日期": latest["日期"], "时间": latest["最晚打卡"]}
            has_early = [r for r in results if r["最早打卡"]]
            if has_early:
                earliest = min(has_early, key=lambda r: r["最早打卡"])
                summary["最早打卡日"] = {"日期": earliest["日期"], "时间": earliest["最早打卡"]}
            return json.dumps({"汇总": summary, "每日明细": results}, ensure_ascii=False, indent=2)
        except Exception as e:
            return f"查询失败: {str(e)[:150]}"

    def _get_today_info(self, args):
        today = date.today()
        ds = today.isoformat()
        info = {"日期": ds, "星期": ["一", "二", "三", "四", "五", "六", "日"][today.weekday()]}
        try:
            from calendar_2026 import get_day_type, get_day_type_label
            info["日期类型"] = get_day_type_label(get_day_type(today))
        except:
            info["日期类型"] = "未知"
        try:
            emp = self._get_empno()
            if emp:
                from user_auth import load_user_config
                from database import query_work_records_range
                from attendance import DayAnalysis
                cfg = load_user_config()
                recs = query_work_records_range(emp, ds, ds)
                a = DayAnalysis(today, recs.get(ds, []), cfg)
                info["打卡次数"] = a.card_count
                info["应打卡次数"] = a.required_cards
                info["加班小时"] = round(a.overtime_hours, 1)
                info["加班费"] = round(a.overtime_pay, 2)
                info["漏打卡"] = a.missed
        except:
            pass
        try:
            ws = json.loads(self._get_water_status({}))
            info["喝水"] = f"{ws['已喝杯数']}/{ws['目标杯数']}杯"
        except:
            pass
        try:
            from todo_manager import todo_mgr
            info["待办"] = len([it for it in todo_mgr.get_all_items() if not it.done])
        except:
            pass
        info["天气"] = "晴 28°C"
        return json.dumps(info, ensure_ascii=False, indent=2)

    def _do_punch_card(self, args):
        ds = args.get("date", date.today().isoformat())
        t = args.get("time", "08:00:00")
        r = args.get("remark", "")
        try:
            from punch_card import generate_sql
            from database import get_connection
            sqls = generate_sql(ds, t, r)
            conn = get_connection()
            c = conn.cursor()
            for sql in sqls:
                c.execute(sql)
            conn.commit()
            conn.close()
            return json.dumps({"success": True, "msg": f"补卡成功 {ds} {t} {r}"}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "msg": str(e)[:150]}, ensure_ascii=False)

    def _get_car_abnormal(self, args):
        today = date.today()
        df = args.get("date_from", today.replace(day=1).isoformat())
        dt = args.get("date_to", today.isoformat())
        try:
            from user_auth import load_user_config
            from car_db import query_car_records
            from calendar_2026 import is_rest_day, get_day_type
            from database import query_work_records_range
            cfg = load_user_config()
            plate = cfg.car_plate if cfg and cfg.car_plate else "粤S0Q780"
            emp = self._get_empno()
            ds_d, de_d = date.fromisoformat(df), date.fromisoformat(dt)
            ym, yme = ds_d.year * 100 + ds_d.month, de_d.year * 100 + de_d.month
            all_rows = []
            while ym <= yme:
                y, m = ym // 100, ym % 100
                try:
                    all_rows.extend(query_car_records(plate, f"{y:04d}{m:02d}"))
                except:
                    pass
                ym = (y + 1) * 100 + 1 if m == 12 else ym + 1
            rows = [r for r in all_rows if df <= str(r.get('ch_crosstime', ''))[:10] <= dt]
            punch_recs = {}
            if emp:
                try:
                    punch_recs = query_work_records_range(emp, df, dt)
                except:
                    pass
            result = []
            for r in rows:
                ct = r.get("ch_crosstime")
                d = ct.date() if hasattr(ct, "date") else ct
                ds_key = d.isoformat() if hasattr(d, "isoformat") else str(d)[:10]
                is_workday = not is_rest_day(d) and get_day_type(d) not in ("holiday", "makeup")
                punch_times = []
                if ds_key in punch_recs:
                    for pr in punch_recs[ds_key]:
                        wt = pr.get("worktime")
                        if wt:
                            punch_times.append(wt.strftime("%H:%M") if hasattr(wt, "strftime") else str(wt)[-5:])
                is_abnormal = True
                if is_workday and punch_times:
                    tp = f"{ct.hour:02d}:{ct.minute:02d}"
                    if "12:05" <= tp <= "13:05":
                        is_abnormal = False
                    elif "17:30" <= tp <= "18:00" and any(t >= "18:00" for t in punch_times):
                        is_abnormal = False
                    elif punch_times and tp > max(punch_times):
                        is_abnormal = False
                elif not is_workday:
                    is_abnormal = False
                if is_abnormal:
                    ts = ct.strftime("%m-%d %H:%M:%S")[:-3] if hasattr(ct, "strftime") else str(ct)
                    result.append({"时间": ts, "进出": "出" if r.get("ch_out", 0) == 1 else "进"})
            return json.dumps(result[:20], ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            return f"查询失败: {str(e)[:150]}"

    def _get_water_status(self, args):
        try:
            from user_auth import load_user_config
            cfg = load_user_config()
            wm = int(cfg.water_ml) if cfg and hasattr(cfg, 'water_ml') else 300
            target = max(1, 2400 // wm)
            try:
                with open(os.path.join(os.path.dirname(__file__), "user_config.json"), "r", encoding="utf-8") as f:
                    ucfg = json.load(f)
                cups = ucfg.get("water_cups_today", 0)
            except:
                cups = 0
            return json.dumps({"已喝杯数": cups, "目标杯数": target, "总毫升": cups * wm}, ensure_ascii=False)
        except:
            return "{}"

    def _query_todos(self, args):
        from todo_manager import todo_mgr
        d = args.get("date")
        only_undone = args.get("only_undone", True)
        items = todo_mgr.get_items(d) if d else todo_mgr.get_all_items()
        result = []
        for it in items:
            if only_undone and it.done:
                continue
            result.append({"日期": it.date, "时间": it.time, "内容": it.content, "完成": it.done, "ID": it.id})
        return json.dumps(result, ensure_ascii=False, indent=2)

    def _add_todo(self, args):
        from todo_manager import todo_mgr
        todo_mgr.add_item(args.get("date", date.today().isoformat()), args.get("time", "09:00"), args.get("content", ""))
        return "✅ 已添加"

    def _toggle_todo(self, args):
        from todo_manager import todo_mgr
        todo_mgr.toggle_item(args.get("date", ""), args.get("id", ""))
        return "✅ 已切换"

    def _delete_todo(self, args):
        from todo_manager import todo_mgr
        todo_mgr.delete_item(args.get("date", ""), args.get("id", ""))
        return "✅ 已删除"

    def _do_water_punch(self, args):
        idx = args.get("index", 0)
        try:
            jpath = os.path.join(os.path.dirname(__file__), "user_config.json")
            if os.path.exists(jpath):
                with open(jpath, "r", encoding="utf-8") as f:
                    ucfg = json.load(f)
            else:
                ucfg = {}
            ucfg["water_cups_today"] = idx + 1
            with open(jpath, "w", encoding="utf-8") as f:
                json.dump(ucfg, f, ensure_ascii=False)
            return f"✅ 第{idx+1}杯水已打卡"
        except:
            return "❌ 打卡失败"

    def _query_car_records(self, args):
        """查询车辆所有出入记录，返回具体时间点"""
        today = date.today()
        df = args.get("date_from", today.replace(day=1).isoformat())
        dt = args.get("date_to", today.isoformat())
        try:
            from user_auth import load_user_config
            from car_db import query_car_records
            cfg = load_user_config()
            plate = cfg.car_plate if cfg and cfg.car_plate else "粤S0Q780"
            ds_d, de_d = date.fromisoformat(df), date.fromisoformat(dt)
            ym, yme = ds_d.year * 100 + ds_d.month, de_d.year * 100 + de_d.month
            all_rows = []
            while ym <= yme:
                y, m = ym // 100, ym % 100
                try:
                    all_rows.extend(query_car_records(plate, f"{y:04d}{m:02d}"))
                except:
                    pass
                ym = (y + 1) * 100 + 1 if m == 12 else ym + 1
            rows = [r for r in all_rows if df <= str(r.get('ch_crosstime', ''))[:10] <= dt]
            result = []
            for r in rows:
                ct = r.get("ch_crosstime")
                ts = ct.strftime("%m-%d %H:%M:%S")[:-3] if hasattr(ct, "strftime") else str(ct)[:14]
                result.append({
                    "时间": ts,
                    "进出": "出" if r.get("ch_out", 0) == 1 else "进",
                    "车牌": r.get("ch_plate", ""),
                })
            import json
            return json.dumps({"记录": result, "总数": len(result), "查询范围": f"{df} ~ {dt}"}, ensure_ascii=False, indent=2)
        except Exception as e:
            return f"查询失败: {str(e)[:150]}"

    def _export_attendance_excel(self, args):
        """生成考勤Excel报表"""
        today = date.today()
        df = args.get("date_from", today.replace(day=1).isoformat())
        dt = args.get("date_to", today.isoformat())
        try:
            import json, openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter

            result_json = self._query_attendance(args)
            data = json.loads(result_json)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "考勤报表"

            hfont = Font(bold=True, size=12, color="FFFFFF")
            hfill = PatternFill(start_color="3370FF", end_color="3370FF", fill_type="solid")
            bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
            ctr = Alignment(horizontal="center", vertical="center")

            summary = data.get("汇总", {})
            ws.merge_cells("A1:F1")
            ws["A1"].value = "哈基米曼波 - 考勤报表 (" + df + " ~ " + dt + ")"
            ws["A1"].font = Font(bold=True, size=14, color="3370FF")
            ws["A1"].alignment = Alignment(horizontal="center")

            row = 3
            for k, v in summary.items():
                ws.cell(row=row, column=1, value=str(k)).font = Font(bold=True)
                ws.cell(row=row, column=2, value=str(v))
                row += 1

            row += 1
            headers = ["日期", "类型", "打卡次数", "加班小时", "加班费", "漏打卡"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = hfont
                cell.fill = hfill
                cell.alignment = ctr
                cell.border = bdr
            row += 1

            for item in data.get("每日明细", []):
                vals = [str(item.get(h, "")) for h in headers]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = bdr
                    cell.alignment = ctr
                row += 1

            filename = "考勤报表_" + df + "_to_" + dt + ".xlsx"
            path = os.path.join(os.path.dirname(__file__), filename)
            wb.save(path)

            return json.dumps({
                "success": True,
                "path": path,
                "filename": filename,
                "说明": "文件已保存到: " + path + "，请打开查看"
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "error": str(e)[:200]}, ensure_ascii=False)


# 全局单例
tool_executor = ToolExecutor()
