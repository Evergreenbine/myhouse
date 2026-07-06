import py_compile

c = open('tools.py', 'r', encoding='utf-8').read()

# 1. Replace stub with real implementation
old_stub = '    def _export_attendance_excel(self, args):\n        return "Excel导出功能在开发中"'

new_impl = '''    def _export_attendance_excel(self, args):
        """生成考勤Excel报表"""
        today = date.today()
        df = args.get("date_from", today.replace(day=1).isoformat())
        dt = args.get("date_to", today.isoformat())
        try:
            import json, openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter

            result_json = self._query_attendance(args)
            data = json.loads(result_json)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "考勤报表"

            hfont = Font(bold=True, size=12, color="FFFFFF")
            hfill = PatternFill(start_color="3370FF", end_color="3370FF", fill_type="solid")
            bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
            ctr = Alignment(horizontal="center", vertical="center")

            summary = data.get("汇总", {})
            ws.merge_cells("A1:F1")
            ws["A1"].value = "哈基米曼波 - 考勤报表 (" + df + " ~ " + dt + ")"
            ws["A1"].font = Font(bold=True, size=14, color="3370FF")
            ws["A1"].alignment = Alignment(horizontal="center")

            row = 3
            for k, v in summary.items():
                ws.cell(row=row, column=1, value=str(k)).font = Font(bold=True)
                ws.cell(row=row, column=2, value=str(v))
                row += 1

            row += 1
            headers = ["日期", "类型", "打卡次数", "加班小时", "加班费", "漏打卡"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = hfont
                cell.fill = hfill
                cell.alignment = ctr
                cell.border = bdr
            row += 1

            for item in data.get("每日明细", []):
                vals = [str(item.get(h, "")) for h in headers]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = bdr
                    cell.alignment = ctr
                row += 1

            filename = "考勤报表_" + df + "_to_" + dt + ".xlsx"
            path = os.path.join(os.path.dirname(__file__), filename)
            wb.save(path)

            return json.dumps({
                "success": True,
                "path": path,
                "filename": filename,
                "说明": "文件已保存到: " + path + "，请打开查看"
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "error": str(e)[:200]}, ensure_ascii=False)'''

c = c.replace(old_stub, new_impl)

# 2. Add tool definition to TOOLS list (before delete_todo)
tool_def = '''    {
        "type": "function",
        "function": {
            "name": "export_attendance_excel",
            "description": (
                "导出考勤报表为Excel文件。生成包含每日明细和汇总的.xlsx文件。\\n"
                "触发词：导出Excel、导出报表、生成Excel、下载考勤、生成报表"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "date_from": {"type": "string", "description": "起始日期 YYYY-MM-DD，默认本月1号"},
                    "date_to": {"type": "string", "description": "结束日期 YYYY-MM-DD，默认今天"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_todo",'''

c = c.replace('''    {
        "type": "function",
        "function": {
            "name": "delete_todo",''', tool_def)

open('tools.py', 'w', encoding='utf-8').write(c)
py_compile.compile('tools.py', doraise=True)
print('tools.py OK')
