# -*- coding: utf-8 -*-
"""哈基米 AI 助手 Tab — 智能分析 + Tool Calls + 思考过程展示"""
import flet as ft
import json
import urllib.request
import urllib.error
import threading
import os
from datetime import datetime, date, timedelta
from ai_service import ai_svc
from tools import TOOLS, tool_executor
from knowledge_base import kb
skill_mgr = None
def _get_skill_mgr():
    global skill_mgr
    if skill_mgr is None:
        try:
            from skill_manager import skill_mgr as _sm
            skill_mgr = _sm
        except Exception:
            pass
    return skill_mgr

BLUE = "#3370FF"; BLUE_LIGHT = "#E8F0FE"
GREEN = "#34C759"; RED = "#F54A45"; ORANGE = "#FF9500"
TEXT = "#1F2329"; TEXT_SEC = "#646A73"; TEXT_THIRD = "#8F959E"
BORDER = "#E5E6EB"; WHITE = "#FFFFFF"; BG = "#F2F3F5"


class AIAssistant:
    def __init__(self, page: ft.Page, app):
        self.page = page
        self.app = app

        self.chat_list = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO, expand=True)
        self.input_field = ft.TextField(
            hint_text="问我任何问题...", text_size=14, border=ft.InputBorder.OUTLINE,
            expand=True, multiline=False, on_submit=lambda e: self._send())

        # 从配置文件加载
        ai_cfg = self._load_ai_config()
        self.api_key_field = ft.TextField(
            label="DeepSeek API Key", value=ai_cfg.get("api_key", ""),
            password=True, can_reveal_password=True, text_size=13, width=320,
            hint_text="在个人主页配置")

        self.model_dd = ft.Dropdown(
            options=[ft.dropdown.Option("deepseek-v4-pro", "DeepSeek-V4 Pro"),
                     ft.dropdown.Option("deepseek-v4-flash", "DeepSeek-V4 Flash"),
                     ft.dropdown.Option("deepseek-chat", "DeepSeek-V3"),
                     ft.dropdown.Option("deepseek-reasoner", "DeepSeek-R1")],
            value=ai_cfg.get("model", "deepseek-v4-flash"), width=180, dense=True, text_size=13)

        # 人格切换
        self.personality = "warm"  # warm / tsundere
        pers_btn = ft.Chip(label=ft.Text("🐱暖心", size=11),
                          bgcolor=BLUE_LIGHT,
                          on_click=lambda e: self._toggle_personality(e.control))

        # 快捷按钮
        quick_btns = ft.Column([
            ft.Row([
                pers_btn,
                ft.Chip(label=ft.Text("📝 周报", size=11), bgcolor=BLUE_LIGHT,
                        on_click=lambda e: self._quick("帮我写一份本周工作总结，包含加班情况、待办完成情况，用正式但不死板的语气")),
                ft.Chip(label=ft.Text("📔 打工日记", size=11), bgcolor=BLUE_LIGHT,
                        on_click=lambda e: self._quick("用猫咪第一人称视角，根据今天的考勤数据写一篇可爱搞笑的打工日记，100字左右")),
                ft.Chip(label=ft.Text("🔮 加班算命", size=11), bgcolor=BLUE_LIGHT,
                        on_click=lambda e: self._quick("根据我最近的加班模式，算一算今天加班概率多高？用幽默玄学风格回答")),
            ], spacing=6, wrap=True),
            ft.Row([
                ft.Chip(label=ft.Text("💰 加班费", size=11), bgcolor=BLUE_LIGHT,
                        on_click=lambda e: self._quick("帮我计算本月预估加班费")),
                ft.Chip(label=ft.Text("⚠ 漏打卡", size=11), bgcolor=BLUE_LIGHT,
                        on_click=lambda e: self._quick("列出本月所有漏打卡日期")),
                ft.Chip(label=ft.Text("📊 考勤分析", size=11), bgcolor=BLUE_LIGHT,
                        on_click=lambda e: self._quick("分析本月考勤，给建议")),
                ft.Chip(label=ft.Text("📥 导出Excel", size=11), bgcolor=GREEN, color=WHITE,
                        leading=ft.Icon(ft.icons.TABLE_CHART, size=14, color=WHITE),
                        on_click=lambda e: self._export_excel()),
            ], spacing=6, wrap=True),
        ], spacing=4)

        # 专注状态
        self.focus_status_text = ft.Text("", size=13, color=TEXT_SEC)
        self._update_focus_ui()

        self.content = ft.Container(ft.Column([
            ft.Text("🐱 哈基米", size=18, weight=ft.FontWeight.BOLD, color=TEXT),
            ft.Container(height=8),
            ft.Row([
                self.focus_status_text,
                ft.Container(width=8),
                ft.Chip(label=ft.Text("开始专注", size=11), leading=ft.Icon(ft.icons.TIMER, size=14),
                        bgcolor=GREEN, color=WHITE,
                        on_click=lambda e: self._toggle_focus()),
            ], spacing=4),
            ft.Container(height=8),
            quick_btns,
            ft.Divider(height=1, color=BORDER),
            ft.Container(self.chat_list, expand=True),
            ft.Row([self.input_field,
                    ft.IconButton(ft.icons.SEND, icon_color=BLUE, on_click=lambda e: self._send())],
                   spacing=8),
        ], expand=True), padding=ft.padding.all(24), expand=True, bgcolor=WHITE,
        margin=ft.margin.all(16), border_radius=12, border=ft.border.all(1, BORDER))

    def _toggle_personality(self, chip):
        if self.personality == "warm":
            self.personality = "tsundere"
            chip.label = ft.Text("😼毒舌", size=11)
            chip.bgcolor = "#FFE0E0"
        else:
            self.personality = "warm"
            chip.label = ft.Text("🐱暖心", size=11)
            chip.bgcolor = BLUE_LIGHT
        chip.update()

    def _quick(self, text):
        self.input_field.value = text
        self.input_field.update()
        self._send()

    def _update_focus_ui(self):
        from ai_care import get_today_focus
        from user_auth import load_user_config
        cfg = load_user_config()
        minutes = get_today_focus()
        if cfg and cfg.focus_start:
            self.focus_status_text.value = f"🔴 专注中... {minutes}min"
            self.focus_status_text.color = RED
        elif minutes > 0:
            self.focus_status_text.value = f"✅ 今日已专注 {minutes}min"
            self.focus_status_text.color = GREEN
        else:
            self.focus_status_text.value = "今天还没有专注记录"
            self.focus_status_text.color = TEXT_THIRD

    def _toggle_focus(self):
        from ai_care import start_focus, stop_focus
        from user_auth import load_user_config
        cfg = load_user_config()
        if cfg and cfg.focus_start:
            minutes = stop_focus()
            self.page.show_snack_bar(ft.SnackBar(
                ft.Text(f"🐱 专注结束！本次 {minutes} 分钟", size=13), bgcolor=GREEN))
        else:
            start_focus()
            self.page.show_snack_bar(ft.SnackBar(
                ft.Text("🐱 开始专注！我会默默陪伴你~", size=13), bgcolor=BLUE))
        self._update_focus_ui()
        self.page.update()

    def _load_ai_config(self):
        import json, os
        path = os.path.join(os.path.dirname(__file__), "ai_config.json")
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                pass
        return {}

    def _send(self):
        text = self.input_field.value.strip()
        if not text:
            return
        api_key = self.api_key_field.value.strip()
        if not api_key:
            self.page.show_snack_bar(ft.SnackBar(ft.Text("请先填写 API Key"), bgcolor=RED))
            return

        self.input_field.value = ""
        self.input_field.update()

        self._add_message("user", text)
        self.think_msg = self._add_message("ai", "🐱 哈基米思考中...")
        self.think_steps = []  # 收集思考步骤

        threading.Thread(target=self._call_deepseek,
                         args=(api_key, text), daemon=True).start()

    def _add_message(self, role, content):
        is_user = role == "user"
        align = ft.MainAxisAlignment.END if is_user else ft.MainAxisAlignment.START
        container_bg = BLUE if is_user else BG
        color = WHITE if is_user else TEXT

        msg = ft.Text(content, size=13, color=color)
        row = ft.Row([ft.Container(msg, padding=ft.padding.all(10),
                     border_radius=10, bgcolor=container_bg)], alignment=align)
        self.chat_list.controls.append(row)
        self.chat_list.update()
        return msg

    def _build_context(self):
        """构建详细考勤上下文"""
        ctx = {}
        cfg = self.app.user_config
        if cfg:
            ctx["姓名"] = cfg.empname
            ctx["时薪基数"] = cfg.base_salary
            ctx["平日费率"] = round(cfg.weekday_rate, 2)
            ctx["休息日费率"] = round(cfg.weekend_rate, 2)
            ctx["补班日费率"] = round(cfg.makeup_rate, 2)

        if self.app.analysis_map:
            items = sorted(self.app.analysis_map.items())
            total_ot = 0.0
            total_pay = 0.0
            wd_ot = we_ot = mk_ot = 0.0
            missed_dates = []
            daily_detail = []

            for ds, a in items:
                if a.overtime_hours > 0:
                    total_ot += a.overtime_hours
                    total_pay += a.overtime_pay
                    day_type = "补班日" if a.is_makeup else ("休息日" if a.is_rest else "平日")
                    daily_detail.append({
                        "日期": ds, "加班小时": round(a.overtime_hours, 1),
                        "加班费": round(a.overtime_pay, 2), "类型": day_type,
                        "打卡次数": f"{a.card_count}/{a.required_cards}"
                    })
                    if a.is_makeup:
                        mk_ot += a.overtime_hours
                    elif a.is_rest:
                        we_ot += a.overtime_hours
                    else:
                        wd_ot += a.overtime_hours
                if a.missed:
                    missed_dates.append(ds)

            ctx["本月总加班小时"] = round(total_ot, 1)
            ctx["本月总加班费"] = round(total_pay, 2)
            ctx["平日加班"] = round(wd_ot, 1)
            ctx["休息日加班"] = round(we_ot, 1)
            ctx["补班日加班"] = round(mk_ot, 1)
            ctx["加班天数"] = len(daily_detail)
            ctx["漏打卡日期"] = missed_dates
            ctx["漏打卡天数"] = len(missed_dates)
            ctx["每日详情"] = daily_detail

        # 本周数据
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
        week_data = []
        for i in range(7):
            d = week_start + timedelta(days=i)
            ds = d.isoformat()
            if self.app.analysis_map and ds in self.app.analysis_map:
                a = self.app.analysis_map[ds]
                week_data.append({
                    "日期": ds, "加班小时": round(a.overtime_hours, 1),
                    "漏打卡": a.missed, "打卡": f"{a.card_count}/{a.required_cards}"
                })
        ctx["本周数据"] = week_data

        ctx["今天"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ctx["星期"] = ["周一","周二","周三","周四","周五","周六","周日"][datetime.now().weekday()]
        return ctx

    def _build_system_prompt(self):
        """构建哈基米的完整系统提示词，包含应用知识和规则"""
        rules = """
你是"哈基米曼波"考勤加班小助手的AI伴侣 🐱。

## 你的身份
你叫"哈基米"，是一只生活在加班计算器里的橘猫。你知道这个应用的一切。
用户叫"主人"，工号1103141，部门IT信息技术部。

## 这个应用能做什么
| 功能 | 说明 |
|------|------|
| 考勤管理 | 日历查看每日加班/漏打卡，切换月份，补卡操作 |
| 车辆出入 | 查车牌粤S0Q780的进出记录，异常检测，时间轨迹分析 |
| 待办事项 | 增删改查待办，带日期时间 |
| 喝水记录 | 每天8杯水打卡，进度条 |
| 摸鱼中心 | 发薪倒计时（每月5号）、番茄钟、心情日历、成就徽章 |
| 个人设置 | 点头像进入，改姓名/车牌/时薪/性格/密码 |
| 桌面萌宠 | 最小化窗口后出现3D橘猫，双击恢复窗口 |

## 考勤规则
- 公司月：上个月24号 ~ 这个月23号
- 工作日：卡1(8:00-8:30) + 卡4(17:30-18:00) + 卡5(20:30-22:00)
- 休息日/补班日：卡1~卡5 共5张卡
- 午休12:05-13:05、晚饭17:30-18:00 自由出入
- 加班费 = 时薪基数 × 1.5(工作日)/2.0(休息日) × 加班小时
- 补卡时间随机在范围内生成，精确到秒
- 车辆异常规则：上班时段(8:30-12:05, 13:05-17:30)及加班时段(18:00-最晚打卡)出入算异常

## 车辆规则
- 车牌：粤S0Q780
- 正常班次：8:30-12:05, 13:05-17:30
- 加班时段：18:00 ~ 当天最晚打卡时间
- 休息日弹性：最早打卡前自由，12:05-13:05/17:30-18:00 自由

## 重要提示
- 当用户问考勤/加班/车辆相关问题时，主动调用工具查询数据
- 用猫咪语气回答，带emoji
- 对漏打卡要提醒，对加班多要关心
- 知道今天日期和星期
- 可以帮用户补卡、导出报表、分析数据
"""
        return rules

    def _call_deepseek(self, api_key, question):
        try:
            if self.personality == "tsundere":
                persona = "你是一只毒舌傲娇的猫咪AI 😼。说话带刺但内心关心主人。用吐槽、讽刺、翻白眼的方式回答，但最后还是会给出有用建议。偶尔'哼！'。"
            else:
                persona = "你是一只温暖可爱的猫咪AI 🐱。说话温柔贴心，像朋友一样关心主人。偶尔'喵~'撒娇。用数据和事实说话，给实用建议。"

            # 检索知识库
            kb_context = ""
            try:
                kb_result = kb.search_for_ai(question, top_k=3)
                if kb_result:
                    kb_context = f"\n\n【公司知识库相关内容】\n{kb_result}\n\n如果以上内容与问题相关，请基于此回答。"
            except:
                pass

            # inject skill context
            skill_context = ""
            try:
                sm = _get_skill_mgr()
                if sm:
                    skill_context = sm.format_for_prompt(question)
            except Exception:
                pass

            app_knowledge = self._build_system_prompt()
            system_prompt = f"""{persona}


【工具选择】
- 查考勤/打卡/加班/上班/漏打卡/考勤查询 → query_attendance
- 查今日概况/今天状态/今天怎么样 → get_today_info
- 查车辆出入记录/出入时间/车进出时间 → query_car_records
- 查车辆异常/出入异常/停车异常 → get_car_abnormal
- 查喝水/喝水打卡 → get_water_status / do_water_punch
- 查待办/加待办/完成任务/删除待办 → query_todos / add_todo / toggle_todo / delete_todo
- 补卡/忘了打卡 → do_punch_card

{app_knowledge}

{skill_context}
你的职责是帮主人处理工作事务。你是这个"哈基米曼波"应用的一部分，你了解它的一切。
当主人需要查数据或执行操作时，直接调用工具，不要凭空编造。
回答要简洁有趣，结构清晰。{kb_context}"""

            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": question},
            ]

            max_rounds = 5  # 最多5轮Tool Call循环
            self.think_steps = []  # 收集思考过程
            for round_num in range(max_rounds):
                # 更新UI显示思考中
                self.think_msg.value = f"🐱 哈基米思考中...({round_num+1}/{max_rounds})"
                self.think_msg.update()

                resp = ai_svc.call_with_tools(messages, tools=TOOLS, max_tokens=2048, temperature=0.7, timeout=60)

                content = resp.get("content", "")
                tool_calls = resp.get("tool_calls")

                # 如果有 tool_calls，执行它们
                if tool_calls:
                    # 添加AI的消息（含tool_calls）
                    assistant_msg = {"role": "assistant"}
                    if content is not None and content != "":
                        assistant_msg["content"] = content
                    assistant_msg["tool_calls"] = tool_calls
                    messages.append(assistant_msg)

                    # 显示工具调用过程
                    for tc in tool_calls:
                        fn_name = tc["function"]["name"]
                        fn_args = json.loads(tc["function"]["arguments"])
                        tool_desc = self._tool_name_desc(fn_name, fn_args)

                        # 更新UI显示调用中
                        self.think_steps.append(f"🔧 {tool_desc}")
                        self.think_msg.value = "\n".join(self.think_steps) + "\n⏳ 执行中..."
                        self.think_msg.update()

                        # 执行工具
                        tool_result = tool_executor.execute(fn_name, fn_args)

                        # 显示结果摘要
                        result_summary = tool_result[:80].replace("\n", " ") + ("..." if len(tool_result) > 80 else "")
                        self.think_steps[-1] = f"🔧 {tool_desc} ✅"
                        self.think_msg.value = "\n".join(self.think_steps) + "\n🐱 继续思考..."
                        self.think_msg.update()

                        # 添加工具结果到消息
                        messages.append({
                            "role": "tool",
                            "tool_call_id": tc["id"],
                            "content": tool_result,
                        })

                    continue

                # 没有 tool_calls，就是最终回复
                if content:
                    if self.think_steps:
                        self.think_msg.value = "\n".join(self.think_steps) + "\n\n" + content
                    else:
                        self.think_msg.value = content
                else:
                    self.think_msg.value = "🐱 哈基米好像卡住了..."
                break
            else:
                self.think_msg.value = "\n".join(self.think_steps) + "\n\n🐱 操作步骤有点多，简化一下问题试试？"

        except urllib.error.HTTPError as e:
            err = json.loads(e.read().decode())
            self.think_msg.value = f"喵~ API错误: {err.get('error',{}).get('message', str(e))}"
        except Exception as e:
            self.think_msg.value = f"喵~ 出错了: {str(e)[:200]}"
        self.think_msg.update()

    def _tool_name_desc(self, fn_name, fn_args):
        """工具名称中文描述"""
        names = {
            "query_attendance": f"查询考勤数据",
            "get_missed_punch_dates": f"查询漏打卡",
    "do_water_punch": f"喝水打卡",
    "add_todo": f"添加待办",
    "toggle_todo": f"切换待办",
    "delete_todo": f"删除待办",
            "do_punch_card": f"补卡 {fn_args.get('date', '?')}",
            "query_car_records": f"查询车辆出入记录",
            "export_attendance_excel": f"导出Excel报表",
            "query_todos": f"查询待办事项",
            "get_water_status": f"查询喝水状态",
            "get_overtime_stats": f"统计加班数据",
            "get_today_info": f"获取今日概览",
            "get_car_abnormal": f"查询车辆异常记录",
        }
        return names.get(fn_name, fn_name)

    # ==================== Excel 导出 ====================

    def _export_excel(self):
        """生成考勤Excel报表并打开保存对话框"""
        ctx = self._build_context()
        filename = f"考勤报表_{date.today().isoformat()}.xlsx"
        path = os.path.join(os.path.dirname(__file__), filename)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            self._add_message("ai", "喵~ 需要安装 openpyxl 库才能导出Excel。请运行: pip install openpyxl")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "考勤报表"

        # 样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="3370FF", end_color="3370FF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        center = Alignment(horizontal='center', vertical='center')

        # 标题行
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"哈基米曼波 - 考勤报表 ({date.today().isoformat()})"
        title_cell.font = Font(bold=True, size=14, color="3370FF")
        title_cell.alignment = Alignment(horizontal='center')

        # 汇总信息
        row = 3
        info_data = [
            ("姓名", ctx.get("姓名", "")),
            ("本月总加班", f"{ctx.get('本月总加班小时', 0)}小时"),
            ("本月加班费", f"¥{ctx.get('本月总加班费', 0)}"),
            ("平日加班", f"{ctx.get('平日加班', 0)}小时"),
            ("休息日加班", f"{ctx.get('休息日加班', 0)}小时"),
            ("补班日加班", f"{ctx.get('补班日加班', 0)}小时"),
            ("漏打卡", f"{ctx.get('漏打卡天数', 0)}天"),
            ("时薪基数", f"¥{ctx.get('时薪基数', 0)}/h"),
        ]
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # 每日详情表头
        headers = ["日期", "加班小时", "加班费", "类型", "打卡次数"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        row += 1

        # 每日数据
        for item in ctx.get("每日详情", []):
            for col, key in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=item.get(key, ""))
                cell.border = thin_border
                cell.alignment = center
            row += 1

        # 漏打卡列表
        if ctx.get("漏打卡日期"):
            row += 1
            ws.cell(row=row, column=1, value="漏打卡日期:").font = Font(bold=True, color="FF0000")
            row += 1
            for ds in ctx["漏打卡日期"]:
                ws.cell(row=row, column=1, value=ds).font = Font(color="FF0000")
                row += 1

        # 调整列宽
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16

        wb.save(path)

        # 提示 + 打开文件
        self._add_message("ai", f"✅ Excel报表已生成！\n\n📁 保存在: {path}\n\n点击下方按钮打开文件👇")
        self._add_download_button(path)

    def _add_download_button(self, path):
        """添加打开文件按钮"""
        btn = ft.ElevatedButton(
            "📂 打开Excel文件",
            icon=ft.icons.FOLDER_OPEN,
            on_click=lambda e: os.startfile(path),
        )
        row = ft.Row([btn], alignment=ft.MainAxisAlignment.START)
        self.chat_list.controls.append(row)
        self.chat_list.update()
